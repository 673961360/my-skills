#!/usr/bin/env python3
"""导出当日原始数据到 Excel（多 sheet）。

用法：
    uv run python scripts/export_data.py              # 当日，prod
    uv run python scripts/export_data.py --env test   # 当日，test
    uv run python scripts/export_data.py --date 20260619  # 指定日期
    uv run python scripts/export_data.py --output my.xlsx  # 指定输出路径

输出：output/全部数据_YYYYMMDD.xlsx
"""

import argparse
import os
import sys
from pathlib import Path

PROJECT_DIR = Path(__file__).resolve().parent
sys.path.insert(0, str(PROJECT_DIR))

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from api_client import get_config
from data_collector import (
    collect_all,
    _parse_date,
    _fmt_display_date,
)


def build_workbook(data: dict) -> Workbook:
    """从 collect_all 返回的 data dict 构建多 sheet Excel。"""

    sheets: list[tuple[str, list[dict]]] = [
        ("O32指令",   data.get("_instructions", [])),
        ("应急回购",   data.get("_emergency_rows", [])),
        ("回购行情",   data.get("_repo_rates", [])),
        ("资金事件",   data.get("_fund_events", [])),
        ("QT日评",     data.get("qt_commentary", {}).get("reports", [])),
    ]

    non_empty = [(n, r) for n, r in sheets if r]
    if not non_empty:
        raise ValueError("所有数据源均为空，无法生成 Excel")

    wb = Workbook()
    wb.remove(wb.active)

    header_font = Font(name="微软雅黑", bold=True, size=10)
    header_fill = PatternFill("solid", fgColor="D9E1F2")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for sheet_name, rows in non_empty:
        ws = wb.create_sheet(title=sheet_name[:31])
        keys = list(rows[0].keys())

        # 表头
        for col_idx, key in enumerate(keys, 1):
            cell = ws.cell(row=1, column=col_idx, value=key)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align

        # 数据行
        for row_idx, record in enumerate(rows, 1):
            for col_idx, key in enumerate(keys, 1):
                ws.cell(row=row_idx + 1, column=col_idx, value=record.get(key))

        # 列宽（采样前 100 行）
        for col_idx, key in enumerate(keys, 1):
            max_len = len(str(key)) * 2
            for row_idx in range(min(100, len(rows))):
                val = rows[row_idx].get(key)
                if val is not None:
                    max_len = max(max_len, len(str(val)) * 2)
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max_len + 2, 40)

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

        # 数字格式
        for row_idx in range(2, len(rows) + 2):
            for col_idx in range(1, len(keys) + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                v = cell.value
                if isinstance(v, float):
                    cell.number_format = '#,##0.00'
                elif isinstance(v, int) and abs(v) > 100000:
                    cell.number_format = '#,##0'

    return wb


def parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser(description="导出当日原始数据到 Excel")
    p.add_argument("--date", type=str, default=None, help="查询日期 YYYYMMDD（默认当日）")
    p.add_argument("--output", type=str, default=None, help="输出路径（默认 output/全部数据_YYYYMMDD.xlsx）")
    p.add_argument("--env", type=str, default="prod",
                   choices=["test", "prod"], help="目标环境（默认 prod）")
    p.add_argument("--use-cache", action="store_true", default=False,
                   help="使用缓存的 API 响应（调试模式）")
    return p.parse_args()


def main():
    args = parse_args()
    os.environ["DTR_ENV"] = args.env
    if args.use_cache:
        os.environ["DTR_USE_CACHE"] = "true"

    query_date = _parse_date(args.date)
    display_date = _fmt_display_date(query_date)

    print(f"日期: {display_date}  环境: {args.env}  缓存: {'ON' if args.use_cache else 'OFF'}")
    print("采集数据...")
    data = collect_all(query_date)

    counts = {
        "O32指令": len(data.get("_instructions", [])),
        "应急回购": len(data.get("_emergency_rows", [])),
        "回购行情": len(data.get("_repo_rates", [])),
        "资金事件": len(data.get("_fund_events", [])),
        "QT日评":   data.get("qt_commentary", {}).get("total", 0),
    }
    for name, n in counts.items():
        print(f"  {name}: {n}")

    wb = build_workbook(data)

    if args.output:
        out = Path(args.output)
    else:
        out = PROJECT_DIR.parent / "output" / f"全部数据_{query_date}.xlsx"
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(out))
    print(f"已保存: {out}  ({out.stat().st_size / 1024:.1f} KB)")


if __name__ == "__main__":
    main()
